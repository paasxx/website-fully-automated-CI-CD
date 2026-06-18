import io
import re
from datetime import datetime
from decimal import Decimal

import msoffcrypto
import openpyxl

from .base import StatementParser, TransactionDTO

INSTALLMENT_RE = re.compile(r"\((\d+)/(\d+)\)\s*$")


class BTGParser(StatementParser):
    BANK = "btg"

    @classmethod
    def detect(cls, headers: set) -> bool:
        return False  # BTG é XLSX, nunca detectado por headers CSV

    def parse(self, file, password=None) -> list[TransactionDTO]:
        raw = file.read()
        buf = io.BytesIO(raw)

        if password:
            decrypted = io.BytesIO()
            office_file = msoffcrypto.OfficeFile(buf)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            wb = openpyxl.load_workbook(decrypted)
        else:
            wb = openpyxl.load_workbook(buf)

        ws = wb.active

        # Localiza o header dinamicamente (linha com 'Data' na col 1, 'Descrição' na col 2)
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if row[1] == 'Data' and row[2] == 'Descrição':
                header_row_idx = i + 1  # 1-based
                break

        if header_row_idx is None:
            raise ValueError("BTG: formato não reconhecido, header não encontrado")

        transactions = []
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            date_val         = row[1]
            description      = row[2]
            amount_val       = row[4]
            transaction_type = row[5]

            if not isinstance(date_val, datetime) or not isinstance(amount_val, (int, float)):
                continue

            amount = Decimal(str(amount_val))
            match = INSTALLMENT_RE.search(description or "")
            clean_desc = INSTALLMENT_RE.sub("", description or "").strip()

            transactions.append(TransactionDTO(
                date=date_val.date(),
                description=clean_desc,
                amount=amount,
                bank=self.BANK,
                is_credit=amount < 0,
                is_installment=bool(match),
                installment_number=int(match.group(1)) if match else None,
                installment_total=int(match.group(2)) if match else None,
                transaction_type=transaction_type,
            ))

        return transactions
